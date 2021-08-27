import tkinter.font as tkFont
import tkinter as tk
from image import *
from tkinter import ttk
from pay import *
import openpyxl


workbook = openpyxl.load_workbook("database.xlsx")
customers = workbook["Sheet1"]
food = workbook["Sheet2"]
food_scratch = workbook["Sheet3"]
gDishes = [["清汤(20元)", "滋补(40元)", "鸳鸯(60元)"],  # 锅底
           ["香菜(10元)", "麻酱(20元)", "韭花(20元)"],  # 佐料
           ["羊肉(30元)", "肥牛(40元)", "白菜(10元)", "茼蒿(20元)"]]  # 菜品


def reset_sheet3():
    global food_scratch
    food_scratch.cell(2, 1, int(food.cell(row=2, column=1).value))
    food_scratch.cell(2, 2, int(food.cell(row=2, column=2).value))
    food_scratch.cell(2, 3, int(food.cell(row=2, column=3).value))
    food_scratch.cell(2, 4, int(food.cell(row=2, column=4).value))
    workbook.save('database.xlsx')


def add_to_listbox(listbox, lst):
    for x in lst:
        listbox.insert(tk.END, x)  # 将x添加到列表框尾部


def do_discount():
    gWin.lbHint["text"] = "饭菜总价：" + str(float(gWin.totalCost * gWin.discount)) + "元"
    gWin.lbHint["fg"] = "black"


def category_changed(event):  # gWin.cbxCategory选项变化时被调用
    gWin.lsbDishes.delete(0, tk.END)  # 删除全部内容,delete(x,y)删除第x项到第y项
    idx = gWin.cbxCategory.current()  # gWin.cbxCategory当前选中的是第idx项
    add_to_listbox(gWin.lsbDishes, gDishes[idx])  # 装入相应菜单
    gWin.lsbDishes.select_set(0, 0)


def bt_delete_click():
    sel = gWin.lsbTable.curselection()
    if sel == ():
        gWin.lbHint["text"] = "您还没有选中要删除的菜"
        gWin.lbHint["fg"] = "red"
    else:
        for i in sel:
            dish = gWin.lsbTable.get(i)
            price = int(dish[7:9])
            price *= int(dish[dish.index("×") + 1:])
            gWin.totalCost -= price
            if dish[4:6] == "羊肉":
                food_scratch.cell(row=2, column=1).value += int(dish[13:])
                workbook.save('database.xlsx')
                bt_add["state"] = tk.NORMAL
                bt_add["text"] = "添加"
                bt_add["fg"] = "black"
                bt_add["font"] = ("楷体", 16, "bold")
            if dish[4:6] == "肥牛":
                food_scratch.cell(row=2, column=2).value += int(dish[13:])
                workbook.save('database.xlsx')
                bt_add["state"] = tk.NORMAL
                bt_add["text"] = "添加"
                bt_add["fg"] = "black"
                bt_add["font"] = ("楷体", 16, "bold")
            if dish[4:6] == "白菜":
                food_scratch.cell(row=2, column=3).value += int(dish[13:])
                workbook.save('database.xlsx')
                bt_add["state"] = tk.NORMAL
                bt_add["text"] = "添加"
                bt_add["fg"] = "black"
                bt_add["font"] = ("楷体", 16, "bold")
            if dish[4:6] == "茼蒿":
                food_scratch.cell(row=2, column=4).value += int(dish[13:])
                workbook.save('database.xlsx')
                bt_add["state"] = tk.NORMAL
                bt_add["text"] = "添加"
                bt_add["fg"] = "black"
                bt_add["font"] = ("楷体", 16, "bold")
        gWin.lbHint["text"] = "饭菜总价：" + str(int(gWin.totalCost * gWin.discount)) + "元"
        gWin.lbHint["fg"] = "black"
        for i in sel[::-1]:
            gWin.lsbTable.delete(i)


def password_click():  # “显示密码”单选框的事件响应函数，点击该单选框时被调用
    if show_password.get():  # showPassword是和cbPassword绑定的tkinter布尔变量
        et_password["show"] = ""  # 使得密码输入框能正常显示密码。Entry有show属性
    else:
        et_password["show"] = "*"  # 使得密码输入框只显示'*'字符


def bt_administrator_click():
    if username.get() == "ad" and password.get() == "123":
        password.set("")  # 将密码输入框清空
        username.set("")  # 将用户名输入框清空
        manage()
    else:
        password.set("")
        username.set("")
        lb_hint["fg"] = "red"
        lb_hint["text"] = "用户名或密码错误，请重新输入!"


def manage():
    lb_welcome.place_forget()
    lb_hint.place_forget()
    lb_username.place_forget()
    lb_password.place_forget()
    et_password.place_forget()
    et_username.place_forget()
    bt_login.place_forget()
    bt_administrator.place_forget()
    cb_password.place_forget()
    bt_register.place_forget()

    gWin.sets = cut_image("background.png", 300, 60, 740, 160)
    gWin.set = tk.Label(gWin, text="请设置菜品份数：", font=('思源黑体', 25, 'bold'), image=gWin.sets, compound="center")
    gWin.set.place(x=300, y=60, width=440, height=100)

    global mutton, cattle, cabbage, lettuce
    mutton, cattle, cabbage, lettuce = tk.StringVar(), tk.StringVar(), tk.StringVar(), tk.StringVar()

    gWin.bg_mutton = cut_image("background.png", 125, 200, 325, 250)
    gWin.lb_mutton = tk.Label(gWin, text="羊肉（当前" + str(food.cell(row=2, column=1).value) + "份）：",
                              font=('思源黑体', 14, 'bold'), image=gWin.bg_mutton, compound="center")
    gWin.lb_mutton.place(x=125, y=200, width=200, height=50)

    gWin.et_mutton = tk.Entry(gWin, textvariable=mutton, font=('', 16, ''))
    gWin.et_mutton.place(x=326, y=210, width=150, height=30)

    gWin.bg_cattle = cut_image("background.png", 515, 200, 715, 250)
    gWin.lb_cattle = tk.Label(gWin, text="肥牛（当前" + str(food.cell(row=2, column=1).value) + "份）：",
                              font=('思源黑体', 14, 'bold'), image=gWin.bg_cattle, compound="center")
    gWin.lb_cattle.place(x=515, y=200, width=200, height=50)

    gWin.et_cattle = tk.Entry(gWin, textvariable=cattle, font=('', 16, ''))
    gWin.et_cattle.place(x=716, y=210, width=150, height=30)

    gWin.bg_cabbage = cut_image("background.png", 125, 320, 325, 370)
    gWin.lb_cabbage = tk.Label(gWin, text="白菜（当前" + str(food.cell(row=2, column=1).value) + "份）：",
                               font=('思源黑体', 14, 'bold'), image=gWin.bg_cabbage, compound="center")
    gWin.lb_cabbage.place(x=125, y=320, width=200, height=50)

    gWin.et_cabbage = tk.Entry(gWin, textvariable=cabbage, font=('', 16, ''))
    gWin.et_cabbage.place(x=326, y=330, width=150, height=30)

    gWin.bg_lettuce = cut_image("background.png", 515, 320, 715, 370)
    gWin.lb_lettuce = tk.Label(gWin, text="茼蒿（当前" + str(food.cell(row=2, column=1).value) + "份）：",
                               font=('思源黑体', 14, 'bold'), image=gWin.bg_lettuce, compound="center")
    gWin.lb_lettuce.place(x=515, y=320, width=200, height=50)

    gWin.et_lettuce = tk.Entry(gWin, textvariable=lettuce, font=('', 16, ''))
    gWin.et_lettuce.place(x=716, y=330, width=150, height=30)

    gWin.bt_sure1 = tk.Button(gWin, text="确定", font=('', 14, 'bold'), command=bt_sure1_click,
                              image=bt, compound="center")
    gWin.bt_sure1.place(x=340, y=460, width=100, height=60)

    gWin.bt_back2 = tk.Button(gWin, text="返回首页", font=('', 14, 'bold'), command=bt_back2_click,
                              image=bt, compound="center")
    gWin.bt_back2.place(x=600, y=460, width=100, height=60)


def bt_sure1_click():
    food.cell(2, 1, mutton.get())
    food.cell(2, 2, cattle.get())
    food.cell(2, 3, cabbage.get())
    food.cell(2, 4, lettuce.get())
    workbook.save('database.xlsx')

    global sub_win
    sub_win = tk.Toplevel(gWin)
    sub_win.title("系统提示")
    sub_win.geometry("250x250+395+175")
    sub_win.configure(bg="white")
    sub_win.resizable(0, 0)
    ht = tk.Label(sub_win, text="设置成功！", bg="white", font=('', 14, ''), relief="flat")
    ht.place(x=75, y=50, width=100, height=30)
    button = tk.Button(sub_win, text="确定", font=('', 14, ''), bg="white", command=button_click)
    button.place(x=75, y=150, width=100, height=40)


def button_click():
    sub_win.destroy()
    gWin.lb_mutton["text"] = "羊肉（当前" + str(food.cell(row=2, column=1).value) + "份）："
    gWin.lb_cattle["text"] = "肥牛（当前" + str(food.cell(row=2, column=2).value) + "份）："
    gWin.lb_cabbage["text"] = "白菜（当前" + str(food.cell(row=2, column=3).value) + "份）："
    gWin.lb_lettuce["text"] = "茼蒿（当前" + str(food.cell(row=2, column=4).value) + "份）："


def bt_back2_click():
    gWin.set.place_forget()
    gWin.lb_mutton.place_forget()
    gWin.et_mutton.place_forget()
    gWin.lb_cattle.place_forget()
    gWin.et_cattle.place_forget()
    gWin.lb_cabbage.place_forget()
    gWin.et_cabbage.place_forget()
    gWin.lb_lettuce.place_forget()
    gWin.et_lettuce.place_forget()
    gWin.bt_sure1.place_forget()
    gWin.bt_back2.place_forget()

    lb_welcome.place(x=170, y=30, width=700, height=90)
    lb_hint.place(x=325, y=150, width=390, height=30)
    lb_hint["text"] = " 请 输 入 您 的 用 户 名 和 密 码 ："
    lb_hint["fg"] = "black"
    lb_username.place(x=320, y=220, width=150, height=30)
    lb_password.place(x=320, y=290, width=150, height=30)
    et_username.place(x=520, y=220, width=200, height=30)
    et_password.place(x=520, y=290, width=200, height=30)
    cb_password.place(x=373, y=350, width=100, height=20)
    bt_administrator.place(x=1, y=1, width=8, height=8)
    bt_login.place(x=340, y=460, width=100, height=60)
    bt_register.place(x=600, y=460, width=100, height=60)


def bt_login_click():
    flag = 0
    for i in range(2, customers.max_row + 1):
        if str(customers.cell(row=i, column=1).value) == username.get() and \
                str(customers.cell(row=i, column=2).value) == password.get():
            flag = 1
            gWin.status = customers.cell(row=i, column=5).value
            gWin.discount = float(customers.cell(row=i, column=3).value)
            gWin.account = float(customers.cell(row=i, column=4).value)
            gWin.sequence = i
            break
    if flag:
        password.set("")
        username.set("")
        login()
    else:
        password.set("")
        username.set("")
        lb_hint["fg"] = "red"
        lb_hint["text"] = "用户名或密码错误，请重新输入!"


def login():
    lb_welcome.place_forget()
    lb_hint.place_forget()
    lb_username.place_forget()
    lb_password.place_forget()
    et_password.place_forget()
    et_username.place_forget()
    bt_login.place_forget()
    bt_administrator.place_forget()
    cb_password.place_forget()
    bt_register.place_forget()

    reset_sheet3()

    big_font = tkFont.Font(family="华文新魏", size=20)
    gWin.option_add("*TCombobox*Listbox*Font", big_font)
    gWin.category = tk.StringVar()  # 对应组合框gWin.cbxCategory收起状态显示的文字
    gWin.cbxCategory = ttk.Combobox(gWin, textvariable=gWin.category, font=big_font)
    gWin.cbxCategory["values"] = ("锅底", "佐料", "菜品")  # 下拉时显示的表象
    gWin.cbxCategory["state"] = "readonly"  # 将gWin.cbxCategory设置为不可输入，只能选择
    gWin.cbxCategory.current(0)  # 选中第0项
    gWin.cbxCategory.place(x=70, y=70, width=300, height=50)

    gWin.lsbDishes = tk.Listbox(gWin, selectmode=tk.SINGLE, exportselection=False, font=("华文新魏", 16, ""))
    # exportselection使得列表框失去输入焦点也能保持选中项目
    gWin.lsbDishes.bind("<Double-Button-1>", lambda e: bt_add_click())
    gWin.lsbDishes.bind("<<ListboxSelect>>", lambda e: gWin.dish_num.set("1"))
    add_to_listbox(gWin.lsbDishes, gDishes[0])  # 装入锅底菜单
    gWin.lsbDishes.select_set(0, 0)  # select_set(x,y)可以选中第x项到第y项(包括y)
    gWin.lsbDishes.place(x=70, y=125, width=300, height=325)

    gWin.cbxCategory.bind("<<ComboboxSelected>>", category_changed)
    # 当组合框下拉后有表现被选中时，会发生ComboboxSelected事件。
    # 此处指定该事件发生时，会调用gWin.categoryChanged函数
    # 指定"<<ComboboxSelected>>"事件的响应函数是gWin.categoryChanged
    gWin.lsbTable = tk.Listbox(gWin, selectmode=tk.EXTENDED, exportselection=False,
                               font=("华文新魏", 20, ""), relief="flat")
    gWin.lsbTable.place(x=635, y=125, width=300, height=325)

    gWin.desk = cut_image("background.png", 635, 60, 955, 115)
    gWin.lb_table = tk.Label(gWin, text="我 的 餐 桌", font=("楷体", 20, "bold"), image=gWin.desk, compound="center")
    gWin.lb_table.place(x=635, y=60, width=320, height=55)

    gWin.id = cut_image("background.png", 70, 520, 230, 545)
    gWin.lb_id = tk.Label(gWin, text="您的账户类型为：", font=("", 14, ""), image=gWin.id, compound="center")
    gWin.lb_id.place(x=70, y=520, width=160, height=25)

    gWin.Id = cut_image("background.png", 230, 520, 310, 545)
    gWin.lbId = tk.Label(gWin, text=gWin.status, fg="red", font=("", 14, "bold"), image=gWin.Id, compound="center")
    gWin.lbId.place(x=230, y=520, width=80, height=25)

    gWin.lbHint = tk.Label(gWin, text="饭菜总价：0元", bg="white", font=("", 16, ""))
    gWin.lbHint.place(x=635, y=451)

    global scrollbar, bt_add, bt_delete, bt_back1, bt_buy
    scrollbar = tk.Scrollbar(gWin, width=20, orient="vertical", command=gWin.lsbTable.yview)
    gWin.lsbTable.configure(yscrollcommand=scrollbar.set)  # 绑定listbox和scrollbar
    scrollbar.place(x=935, y=125, width=20, height=325)

    gWin.num = cut_image("background.png", 407, 80, 477, 110)
    gWin.number = tk.Label(gWin, text="数量：", font=("", 16, "bold"), image=gWin.num, compound="center")
    gWin.number.place(x=407, y=80, width=70, height=30)

    gWin.dish_num = tk.StringVar(value="1")
    gWin.sp_num = tk.Spinbox(gWin, width=70, from_=1, to=1000, textvariable=gWin.dish_num, font=("", 16, ""))
    gWin.sp_num.place(x=487, y=80, width=100, height=30)

    global bt1
    bt1 = get_image("button.png", 120, 50)
    bt_add = tk.Button(gWin, text="添 加", font=("楷体", 16, "bold"), command=bt_add_click, image=bt1, compound="center")
    bt_add.place(x=442, y=150, width=120, height=50)

    bt_delete = tk.Button(gWin, text="删 除", font=("楷体", 16, "bold"), command=bt_delete_click,
                          image=bt1, compound="center")
    bt_delete.place(x=442, y=230, width=120, height=50)

    bt_back1 = tk.Button(gWin, text="返回首页", font=("楷体", 16, "bold"), command=bt_back1_click,
                         image=bt1, compound="center")
    bt_back1.place(x=442, y=310, width=120, height=50)

    bt_buy = tk.Button(gWin, text="下 单", font=("楷体", 16, "bold"), command=bt_buy_click,
                       image=bt1, compound="center")
    bt_buy.place(x=442, y=390, width=120, height=50)

    gWin.bt_VIP = tk.Button(gWin, text="开通会员", font=("楷体", 16, "bold"), command=bt_vip_click,
                            image=bt1, compound="center")
    if gWin.discount == 1:
        gWin.bt_VIP.place(x=442, y=470, width=120, height=50)


def bt_vip_click():
    global sub_vip, code
    sub_vip = tk.Toplevel(gWin)
    sub_vip.title("开通会员")
    sub_vip.geometry("800x500+380+100")
    sub_vip.configure(bg="white")
    sub_vip.resizable(0, 0)
    lb_vip = tk.Label(sub_vip, text="     开通会员需付款200元，其中的150元将会存入火锅店账户余额中使用。" + '\n' + '\n' +
                                    "会员用户可享受8折优惠哦！",
                      font=("", 16, "bold"), bg="white", anchor="w")
    lb_vip.place(x=8, y=50, width=780)

    code = get_image("code.jpg", 280, 280)
    lb_code = tk.Label(sub_vip, image=code, compound="center")
    lb_code.place(x=50, y=180)
    bt_ok = tk.Button(sub_vip, text="支付完成", font=('', 20, ''), bg="white", command=bt_ok_click)
    bt_ok.place(x=500, y=300, width=150, height=60)


def bt_ok_click():
    sub_vip.destroy()
    customers.cell(gWin.sequence, 5, "会员账户")
    customers.cell(gWin.sequence, 3, 0.8)
    customers.cell(gWin.sequence, 4, 150)
    workbook.save('database.xlsx')
    gWin.status = customers.cell(row=gWin.sequence, column=5).value
    gWin.lbId["text"] = gWin.status
    gWin.discount = customers.cell(row=gWin.sequence, column=3).value


def bt_buy_click():
    global pay_set, font_size, pfa, PAD, pay, entry_pay_from_account
    pay_set = tk.Toplevel(gWin)
    pay_set.title("下单")
    pay_set.resizable(0, 0)
    pfa_cmd = pay_set.register(test)
    pay = float(gWin.totalCost * gWin.discount)
    pfa = tk.StringVar()
    pfa.set(min(pay, gWin.account))

    font_size = 20
    entry_pay_from_account = tk.Entry(pay_set, textvariable=pfa, validate='key', validatecommand=(pfa_cmd, '%P'))
    label_pay = tk.Label(pay_set, text="本次消费需支付：" + str(pay) + "元",
                         fg="black", font=('华文行楷', font_size, 'bold'))
    label_account = tk.Label(pay_set, text="账户余额：" + str(gWin.account) + "元",
                             fg="black", font=('华文行楷', font_size, 'bold'))
    label_pay_from_account = tk.Label(pay_set, text="请输入从账户中支付：", fg="black", font=('华文行楷', font_size, 'bold'))
    label_yuan = tk.Label(pay_set, text="元", fg="black", font=('华文行楷', font_size, 'bold'))

    bt_sure = tk.Button(pay_set, text="支付", fg="black", font=('华文行楷', font_size, 'bold'), bg="white",
                        command=pay_interface)
    bt_back = tk.Button(pay_set, text="返回", fg="black", font=('华文行楷', font_size, 'bold'), bg="white",
                        command=pay_set.destroy)

    COL = 2
    PAD = 10
    label_pay.grid(row=0, column=0, columnspan=COL + 1, padx=PAD, pady=PAD)
    label_account.grid(row=1, column=0, columnspan=COL + 1, padx=PAD, pady=PAD)
    label_pay_from_account.grid(row=2, column=0, padx=PAD, pady=PAD)
    label_yuan.grid(row=2, column=2, padx=PAD, pady=PAD)
    entry_pay_from_account.grid(row=2, column=1, padx=PAD, pady=PAD)
    bt_sure.grid(row=4, column=0, padx=PAD, pady=PAD)
    bt_back.grid(row=4, column=1, padx=PAD, pady=PAD)


def pay_interface():  # 支付界面
    global pay_face
    pay_face = tk.Toplevel(pay_set)
    pay_face.resizable(0, 0)
    pay_from_account = tk.Label(pay_face, text="已从账户中支付：" + str(pfa.get()) + "元",
                                fg="black", font=('华文行楷', font_size, 'bold'))
    pay_back = tk.Button(pay_face, text="返回", fg="black", font=('华文行楷', font_size, 'bold'), bg="white",
                         command=pay_face.destroy)
    pay_over = tk.Button(pay_face, text="支付成功", fg="black", font=('华文行楷', font_size, 'bold'), bg="white",
                         command=pay_over_click)
    pfm = pay - float(pfa.get())  # 需另外支付
    if float(pfa.get()) <= gWin.account:
        if pfm > 0:
            pay_pay_from_money = tk.Label(pay_face, text="还需支付：" + str(pfm) + "元",
                                          fg="black", font=('华文行楷', font_size, 'bold'))
            pay_qrcode = tk.Label(pay_face, text="扫描下方二维码进行支付", fg="black", font=('华文行楷', font_size, 'bold'))

            global code1
            code1 = get_image("code.jpg", 180, 180)
            pay_code_jpg = tk.Label(pay_face, image=code1, compound="center")
            pay_from_account.grid(row=0, column=0, columnspan=2, padx=PAD, pady=PAD)
            pay_pay_from_money.grid(row=1, column=0, columnspan=2, padx=PAD, pady=PAD)
            pay_qrcode.grid(row=2, column=0, columnspan=2, padx=PAD, pady=PAD)
            pay_code_jpg.grid(row=3, column=0, columnspan=2, padx=PAD, pady=PAD)
            pay_over.grid(row=4, column=0, padx=PAD, pady=PAD)
            pay_back.grid(row=4, column=1, padx=PAD, pady=PAD)
        elif pfm == 0:
            pay_from_account.grid(row=0, column=0, padx=PAD, pady=PAD)
            pay_over.grid(row=1, column=0, padx=PAD, pady=PAD)
    elif float(pfa.get()) > gWin.totalCost or pfm < 0:
        pay_error = tk.Label(pay_face, text="金额有误，请重新输入！", fg="red", font=('华文行楷', font_size, 'bold'))
        pay_error.grid(row=0, column=0, padx=PAD, pady=PAD)
        pay_back.grid(row=1, column=0, padx=PAD, pady=PAD)


def pay_over_click():
    pay_face.destroy()
    red_packet = tk.Toplevel(pay_set)
    num_packet = (int(Red_envelope(gWin.totalCost, c=0.1)) // 10) * 10 + 8
    label_packet = tk.Label(red_packet, text="恭喜您获得红包 " + str(num_packet) + " 元",
                            fg="black", font=('华文行楷', font_size, 'bold'))
    packet_sure = tk.Button(red_packet, text="确定", fg="black", font=('华文行楷', font_size, 'bold'), bg="white",
                            command=red_packet.destroy)
    label_packet.grid(row=0, column=0, padx=PAD, pady=PAD)
    packet_sure.grid(row=1, column=0, padx=PAD, pady=PAD)
    gWin.account = gWin.account - float(pfa.get()) + num_packet
    customers.cell(gWin.sequence, 4, gWin.account)
    food.cell(2, 1, int(food_scratch.cell(row=2, column=1).value))
    food.cell(2, 2, int(food_scratch.cell(row=2, column=2).value))
    food.cell(2, 3, int(food_scratch.cell(row=2, column=3).value))
    food.cell(2, 4, int(food_scratch.cell(row=2, column=4).value))
    workbook.save('database.xlsx')


def bt_back1_click():
    gWin.cbxCategory.place_forget()
    gWin.lsbDishes.place_forget()
    gWin.lsbTable.place_forget()
    gWin.lb_table.place_forget()
    gWin.lbId.place_forget()
    gWin.lbHint.place_forget()
    scrollbar.place_forget()
    gWin.number.place_forget()
    gWin.sp_num.place_forget()
    bt_add.place_forget()
    bt_delete.place_forget()
    bt_back1.place_forget()
    bt_buy.place_forget()
    gWin.bt_VIP.place_forget()
    gWin.lb_id.place_forget()

    lb_welcome.place(x=170, y=30, width=700, height=90)
    lb_hint["text"] = " 请 输 入 您 的 用 户 名 和 密 码 ："
    lb_hint["fg"] = "black"
    lb_hint.place(x=325, y=150, width=390, height=30)
    lb_username.place(x=320, y=220, width=150, height=30)
    lb_password.place(x=320, y=290, width=150, height=30)
    et_username.place(x=520, y=220, width=200, height=30)
    et_password.place(x=520, y=290, width=200, height=30)
    cb_password.place(x=373, y=350, width=100, height=20)
    bt_administrator.place(x=1, y=1, width=8, height=8)
    bt_login.place(x=340, y=460, width=100, height=60)
    bt_register.place(x=600, y=460, width=100, height=60)


def bt_add_click():
    global food_scratch
    # btAdd["state"] = tk.DISABLED  tk.NORMAL
    sel = gWin.lsbDishes.curselection()  # sel形如 (1,2,3)
    if sel == ():
        gWin.lbHint["text"] = "您还没有选中要添加的菜"
        gWin.lbHint["fg"] = "red"
    else:
        dish = gWin.lsbDishes.get(sel)
        price, num = int(dish[3:5]), gWin.dish_num.get()
        Num = int(num)

        if dish[0:2] == "羊肉":
            bt_add["state"] = tk.NORMAL
            if food_scratch.cell(row=2, column=1).value < Num:
                bt_add["state"] = tk.DISABLED
                bt_add["text"] = "所选数量过多！"
                bt_add["fg"] = "red"
                bt_add["font"] = ("楷体", 10, "")
            else:
                food_scratch.cell(row=2, column=1).value -= Num
                bt_add["state"] = tk.NORMAL
                bt_add["text"] = "添加"
                bt_add["fg"] = "black"
                bt_add["font"] = ("楷体", 16, "bold")
                workbook.save('database.xlsx')
        if dish[0:2] == "肥牛":
            bt_add["state"] = tk.NORMAL
            if food_scratch.cell(row=2, column=2).value < Num:
                bt_add["state"] = tk.DISABLED
                bt_add["text"] = "所选数量过多！"
                bt_add["fg"] = "red"
                bt_add["font"] = ("楷体", 10, "")
            else:
                food_scratch.cell(row=2, column=2).value -= Num
                workbook.save('database.xlsx')
                bt_add["state"] = tk.NORMAL
                bt_add["text"] = "添加"
                bt_add["fg"] = "black"
                bt_add["font"] = ("楷体", 16, "bold")
        if dish[0:2] == "白菜":
            bt_add["state"] = tk.NORMAL
            if food_scratch.cell(row=2, column=3).value < Num:
                bt_add["state"] = tk.DISABLED
                bt_add["text"] = "所选数量过多！"
                bt_add["fg"] = "red"
                bt_add["font"] = ("楷体", 10, "")
            else:
                food_scratch.cell(row=2, column=3).value -= Num
                workbook.save('database.xlsx')
                bt_add["state"] = tk.NORMAL
                bt_add["text"] = "添加"
                bt_add["fg"] = "black"
                bt_add["font"] = ("楷体", 16, "bold")
        if dish[0:2] == "茼蒿":
            bt_add["state"] = tk.NORMAL
            if food_scratch.cell(row=2, column=4).value < Num:
                bt_add["state"] = tk.DISABLED
                bt_add["text"] = "所选数量过多！"
                bt_add["fg"] = "red"
                bt_add["font"] = ("楷体", 10, "")
            else:
                food_scratch.cell(row=2, column=4).value -= Num
                workbook.save('database.xlsx')
                bt_add["state"] = tk.NORMAL
                bt_add["text"] = "添加"
                bt_add["fg"] = "black"
                bt_add["font"] = ("楷体", 16, "bold")

        if bt_add["state"] == tk.NORMAL:
            gWin.lsbTable.insert(tk.END, "[" + gWin.category.get() + "]" + dish + " ×" + num)
            gWin.totalCost += price * Num
            gWin.lbHint["text"] = "饭菜总价：" + str(int(gWin.totalCost * gWin.discount)) + "元"
            gWin.lbHint["fg"] = "black"


def bt_register_click():
    global gWin
    bt_login.place_forget()
    lb_hint.place_forget()
    cb_password.place_forget()
    bt_administrator.place_forget()
    bt_register.place_forget()
    lb_username.place_forget()
    lb_password.place_forget()
    et_password.place_forget()
    et_username.place_forget()

    gWin.repassword, gWin.username1, gWin.password1 = tk.StringVar(), tk.StringVar(), tk.StringVar()

    gWin.name1 = cut_image("background.png", 320, 220, 470, 250)
    gWin.lbUsername1 = tk.Label(gWin, text=" 用 户 名 ：", image=gWin.name1, fg="black", font=('思源黑体', 16, ''),
                                compound="center")
    gWin.lbUsername1.place(x=320, y=220, width=150, height=30)

    gWin.psw1 = cut_image("background.png", 320, 290, 470, 320)
    gWin.lbPassword1 = tk.Label(gWin, text=" 密 码 ：", image=gWin.psw1, fg="black", font=('思源黑体', 16, ''),
                                compound="center")
    gWin.lbPassword1.place(x=320, y=290, width=150, height=30)

    gWin.etUsername1 = tk.Entry(gWin, textvariable=gWin.username1, font=('', 14, ''))
    gWin.etUsername1.place(x=520, y=220, width=200, height=30)

    gWin.etPassword1 = tk.Entry(gWin, textvariable=gWin.password1, show="*", font=('', 14, ''))
    gWin.etPassword1.place(x=520, y=290, width=200, height=30)

    lb_welcome["text"] = "用户注册"
    lb_welcome["fg"] = "black"

    gWin.rpsw = cut_image("background.png", 280, 360, 500, 390)
    gWin.lbrePassword = tk.Label(gWin, text=" 再次输入密码 ：", image=gWin.rpsw, fg="black", font=('思源黑体', 16, ''),
                                 compound="center")
    gWin.lbrePassword.place(x=280, y=360, width=150, height=30)

    gWin.etrePassword = tk.Entry(gWin, textvariable=gWin.repassword, show="*", font=('', 14, ''))
    gWin.etrePassword.place(x=520, y=360, width=200, height=30)

    gWin.bt_register1 = tk.Button(gWin, text="注 册", image=bt, fg="black", font=('', 14, 'bold'),
                                  command=bt_register1_click, compound="center")
    gWin.bt_register1.place(x=340, y=460, width=100, height=60)

    gWin.bt_back3 = tk.Button(gWin, text="返 回", image=bt, fg="black", font=('', 14, 'bold'),
                              command=bt_back3_click, compound="center")
    gWin.bt_back3.place(x=600, y=460, width=100, height=60)


def bt_back3_click():
    bt_login.place(x=340, y=460, width=100, height=60)
    lb_hint.place(x=325, y=150, width=390, height=30)
    bt_register.place(x=600, y=460, width=100, height=60)
    cb_password.place(x=373, y=350, width=100, height=20)
    bt_administrator.place(x=1, y=1, width=8, height=8)
    lb_username.place(x=320, y=220, width=150, height=30)
    lb_password.place(x=320, y=290, width=150, height=30)
    et_username.place(x=520, y=220, width=200, height=30)
    et_password.place(x=520, y=290, width=200, height=30)
    lb_welcome["text"] = "欢迎光临"
    lb_welcome["fg"] = "red"
    gWin.bt_register1.place_forget()
    gWin.lbrePassword.place_forget()
    gWin.etrePassword.place_forget()
    gWin.lbUsername1.place_forget()
    gWin.lbPassword1.place_forget()
    gWin.etUsername1.place_forget()
    gWin.etPassword1.place_forget()
    gWin.bt_back3.place_forget()
    Checktxt.place_forget()


def bt_register1_click():
    global gWin, Checktxt
    if gWin.username1.get() == "" or gWin.password1.get() == "" or gWin.repassword.get() == "":
        gWin.check = cut_image("background.png", 350, 150, 690, 200)
        Checktxt = tk.Label(gWin, text="输入的内容不能为空！", image=gWin.check, fg="red",
                            font=("楷体", 14, "bold"), compound="center")
        Checktxt.place(x=350, y=150, width=340, height=50)
    elif gWin.repassword.get() != gWin.password1.get():
        gWin.password1.set("")
        gWin.repassword.set("")
        gWin.check = cut_image("background.png", 350, 150, 690, 200)
        Checktxt = tk.Label(gWin, text="两次输入密码不一致，请重新输入！", image=gWin.check, fg="red",
                            font=("楷体", 14, "bold"), compound="center")
        Checktxt.place(x=350, y=150, width=340, height=50)
    else:
        flag1 = 1
        for i in range(2, customers.max_row + 1):
            if str(customers.cell(row=i, column=1).value) == gWin.username1.get():
                flag1 = 0
                break
        if flag1:
            row = customers.max_row
            customers.cell(row + 1, 1, gWin.username1.get())
            customers.cell(row + 1, 2, gWin.password1.get())
            customers.cell(row + 1, 3, 1)
            customers.cell(row + 1, 4, 0)
            customers.cell(row + 1, 5, "普通账户")
            workbook.save('database.xlsx')

            global success
            success = tk.Toplevel(gWin)
            success.title("系统提示")
            success.geometry("250x250+395+175")
            success.configure(bg="white")
            success.resizable(0, 0)
            ht1 = tk.Label(success, text="注册成功！", bg="white", font=('', 14, ''), relief="flat")
            ht1.place(x=75, y=50, width=100, height=30)
            button1 = tk.Button(success, text="确定", font=('', 14, ''), bg="white", command=button1_click)
            button1.place(x=75, y=150, width=100, height=40)
        else:
            gWin.username1.set("")
            gWin.password1.set("")
            gWin.repassword.set("")
            gWin.check = cut_image("background.png", 350, 150, 690, 200)
            Checktxt = tk.Label(gWin, text="该用户名已存在，请重新输入！", image=gWin.check, fg="red",
                                font=("楷体", 14, "bold"), compound="center")
            Checktxt.place(x=350, y=150, width=340, height=50)


def button1_click():
    success.destroy()
    bt_login.place(x=340, y=460, width=100, height=60)
    lb_hint.place(x=325, y=150, width=390, height=30)
    bt_register.place(x=600, y=460, width=100, height=60)
    cb_password.place(x=373, y=350, width=100, height=20)
    bt_administrator.place(x=1, y=1, width=8, height=8)
    lb_username.place(x=320, y=220, width=150, height=30)
    lb_password.place(x=320, y=290, width=150, height=30)
    et_username.place(x=520, y=220, width=200, height=30)
    et_password.place(x=520, y=290, width=200, height=30)
    lb_welcome["text"] = "欢迎光临"
    lb_welcome["fg"] = "red"
    gWin.bt_register1.place_forget()
    gWin.lbrePassword.place_forget()
    gWin.etrePassword.place_forget()
    gWin.lbUsername1.place_forget()
    gWin.lbPassword1.place_forget()
    gWin.etUsername1.place_forget()
    gWin.etPassword1.place_forget()
    gWin.bt_back3.place_forget()
    Checktxt.place_forget()


def main():
    global gWin, background, canvas, lb_welcome, lb_hint, username, password, lb_username, lb_password, \
        et_username, et_password, show_password, cb_password, bt_administrator, bt_login, bt_register, bt
    gWin = tk.Tk()
    gWin.title("火锅点菜系统")
    gWin.geometry("1040x600+280+100")
    gWin.resizable(0, 0)

    background = get_image("background.png", 1040, 600)
    canvas = tk.Canvas(gWin, width=1040, height=600)
    canvas.create_image(520, 300, image=background)
    canvas.pack()

    gWin.totalCost, gWin.discount, gWin.account, gWin.sequence, gWin.status = 0, 1, 0, 1, "普通账户"  # 总价、折扣和账户余额

    welcome = get_image("welcome.png", 700, 90)
    lb_welcome = tk.Label(gWin, text="   欢  迎  光  临  ！", fg="red", font=('华文行楷', 50, 'bold'),
                          image=welcome, compound="center")
    lb_welcome.place(x=170, y=30, width=700, height=90)

    hint = cut_image("background.png", 325, 150, 715, 180)
    lb_hint = tk.Label(gWin, image=hint, text=" 请 输 入 您 的 用 户 名 和 密 码 ：", fg="black",
                       font=('黑体', 16, ''), compound="center")
    lb_hint.place(x=325, y=150, width=390, height=30)

    username, password = tk.StringVar(), tk.StringVar()

    name = cut_image("background.png", 320, 220, 470, 250)
    lb_username = tk.Label(gWin, text=" 用 户 名 ：", image=name, fg="black",
                           font=('黑体', 16, 'bold'), compound="center")
    lb_username.place(x=320, y=220, width=150, height=30)

    psw = cut_image("background.png", 320, 290, 470, 320)
    lb_password = tk.Label(gWin, text=" 密 码 ：", image=psw,
                           fg="black", font=('黑体', 16, 'bold'), compound="center")
    lb_password.place(x=320, y=290, width=150, height=30)

    et_username = tk.Entry(gWin, textvariable=username, font=('', 14, ''))
    et_username.place(x=520, y=220, width=200, height=30)

    et_password = tk.Entry(gWin, textvariable=password, show="*", font=('', 14, ''))
    # Entry的属性show="*"表示该输入框不论内容是啥，只显示'*'字符，为""则正常显示
    et_password.place(x=520, y=290, width=200, height=30)

    show_password = tk.BooleanVar()  # 用于关联“显示密码”单选框
    show_password.set(False)  # 使得cb_password开始就是非选中状态
    show = cut_image("background.png", 398, 350, 468, 370)
    cb_password = tk.Checkbutton(gWin, text=" 显示密码", bg="white", fg="black", font=('黑体', 12, ''),
                                 image=show, variable=show_password, command=password_click,
                                 compound="center")
    # cbPassword关联变量showPassword，其事件响应函数是cbPassword_click，即点击它时，会调用 cbPassword_click()
    cb_password.place(x=373, y=350, width=100, height=20)

    administrator = cut_image("background.png", 1, 1, 9, 9)
    bt_administrator = tk.Button(gWin, text=" ", bg="white", command=bt_administrator_click, relief="groove",
                                 image=administrator, compound="center")
    bt_administrator.place(x=1, y=1, width=8, height=8)

    bt = get_image("button.png", 100, 60)
    bt_login = tk.Button(gWin, text="登 录", bg="white", fg="black", font=('', 14, 'bold'), command=bt_login_click,
                         image=bt, compound="center")
    # 点击btLogin按钮会执行btLogin_click()
    bt_login.place(x=340, y=460, width=100, height=60)
    bt_register = tk.Button(gWin, text="注 册", bg="white", fg="black", font=('', 14, 'bold'),
                            command=bt_register_click, image=bt, compound="center")
    bt_register.place(x=600, y=460, width=100, height=60)

    gWin.mainloop()


main()
